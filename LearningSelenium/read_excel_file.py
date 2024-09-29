from openpyxl import workbook,load_workbook

wb = load_workbook(filename='demo.xlsx')
sh=wb.active
print(sh['B1'].value)

# if we want to get a sheet column in a single line
# print(wb['Sheet1']['B2'].value)

print(sh.cell(2,3).value)

print(sh.max_row," ", sh.max_column)